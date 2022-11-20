#@title
import argparse
import json
import os
from pathlib import Path
from threading import Thread

import numpy as np
import torch
import yaml
from tqdm import tqdm
from canvas_DynaPP import *

from models.experimental import attempt_load
from utils.datasets import create_dataloader
from utils.general import coco80_to_coco91_class, check_dataset, check_file, check_img_size, check_requirements, \
    box_iou, non_max_suppression, scale_coords, xyxy2xywh, xywh2xyxy, set_logging, increment_path, colorstr, box_center
from utils.metrics import ap_per_class, ConfusionMatrix
from utils.plots import plot_images, output_to_target, plot_study_txt
from utils.torch_utils import select_device
from utils.torch_utils import time_sync as time_synchronized
import torchvision
from openpyxl import Workbook
from openpyxl import load_workbook

def AUAIR_change_cls(cls):
    change = torch.cuda.LongTensor([0,0,5,1,3,2,5,5,6,4])
    return change[cls.type('torch.cuda.LongTensor')]


def UAVDT_change_cls(cls):
    change = torch.cuda.LongTensor([-1,-1,-1,0,-1,1,-1,-1,2,-1])
    return change[cls.type('torch.cuda.LongTensor')]


def test(data,
         weights=None,
         batch_size=32,
         imgsz=640,
         conf_thres=0.001,
         iou_thres=0.6,  # for NMS
         save_json=False,
         single_cls=False,
         augment=False,
         verbose=False,
         model=None,
         dataloader=None,
         save_dir=Path(''),  # for saving images
         save_txt=False,  # for auto-labelling
         save_hybrid=False,  # for hybrid auto-labelling
         save_conf=False,  # save auto-label confidences
         plots=True,
         wandb_logger=None,
         compute_loss=None,
         half_precision=True,
         is_coco=False,
         opt=None):
    # Initialize/load model and set device  
        
    training = model is not None
    if training:  # called by train.py
        device = next(model.parameters()).device  # get model device

    else:  # called directly
        set_logging()
        device = select_device(opt.device, batch_size=batch_size)

        # Directories
        save_dir = increment_path(Path(opt.project) / opt.name, exist_ok=opt.exist_ok)  # increment run
        (save_dir / 'labels' if save_txt else save_dir).mkdir(parents=True, exist_ok=True)  # make dir

        # Load model
        model = attempt_load(weights, map_location=device)  # load FP32 model
        gs = max(int(model.stride.max()), 32)  # grid size (max stride)
        # imgsz = check_img_size(imgsz, s=gs)  # check img_size

        # Multi-GPU disabled, incompatible with .half() https://github.com/ultralytics/yolov5/issues/99
        # if device.type != 'cpu' and torch.cuda.device_count() > 1:
        #     model = nn.DataParallel(model)

    # Half
    half = device.type != 'cpu' and half_precision  # half precision only supported on CUDA
    if half:
        model.half()

    # Configure
    model.eval()
    if isinstance(data, str):
        is_coco = data.endswith('coco.yaml')
        with open(data) as f:
            data = yaml.safe_load(f)
    check_dataset(data)  # check
    nc = 1 if single_cls else int(data['nc'])  # number of classes
    iouv = torch.linspace(0.5, 0.95, 10).to(device)  # iou vector for mAP@0.5:0.95
    niou = iouv.numel()

    # Logging
    log_imgs = 0
    if wandb_logger and wandb_logger.wandb:
        log_imgs = min(wandb_logger.log_imgs, 100)
    # Dataloader
    if opt.pack == True:
        rect = True
    else:
        rect = False
    if not training:
        if device.type != 'cpu':
            model(torch.zeros(1, 3, imgsz, imgsz).to(device).type_as(next(model.parameters())))  # run once
        task = opt.task if opt.task in ('train','val','test') else 'val'  # path to train/val/test images
        dataloader = create_dataloader(data[task], imgsz, batch_size, gs, single_cls, pad=0, rect=rect,
                                       prefix=colorstr(f'{task}: '))[0]

    seen = 0
    confusion_matrix = ConfusionMatrix(nc=nc)
    names = {k: v for k, v in enumerate(data['names'])}
    coco91class = coco80_to_coco91_class()
    s = ('%20s' + '%12s' * 6) % ('Class', 'Images', 'Labels', 'P', 'R', 'mAP@.5', 'mAP@.5:.95')
    p, r, f1, mp, mr, map50, map, t0, t1,t2,t3, t_pack, pack_t1, pack_t2 = 0., 0., 0., 0., 0., 0., 0., 0., 0. ,0., 0., 0., 0., 0.
    packed_img_h, packed_img_w = [] , []
    all_packed_img = []
    all_anchor_img = []
    loss = torch.zeros(3, device=device)
    jdict, stats, ap, ap_class, wandb_images = [], [], [], [], []
    key_frame = 0
    prev_out = None
    duration = opt.duration
    f = save_dir / 'packed_frames'
    os.mkdir(f)
    
    for batch_i, (img, targets, paths, shapes) in enumerate(tqdm(dataloader, desc=s)):
        img = img.to(device, non_blocking=True)
        img = img.half() if half else img.float()  # uint8 to fp16/32
        img /= 255.0  # 0 - 255 to 0.0 - 1.0
        targets = targets.to(device)
        nb, _, height, width = img.shape  # batch size, channels, height, width

        with torch.no_grad():
            # Run model
            if opt.pack == False or key_frame ==0:
                pack = False
                where_old_image= None
                where_new_image = None
                t = time_synchronized()
                out, train_out = model(img, augment=augment)  # inference and training outputs
                t0 += time_synchronized() - t
                all_anchor_img.append(img.size(2)*img.size(3))
                # Compute loss
                if compute_loss:
                    loss += compute_loss([x.float() for x in train_out], targets)[1][:3]  # box, obj, cls

                # Run NMS
                targets[:, 2:] *= torch.Tensor([width, height, width, height]).to(device)  # to pixels
                lb = [targets[targets[:, 0] == i, 1:] for i in range(nb)] if save_hybrid else []  # for autolabelling
                t = time_synchronized()
                out = non_max_suppression(out, conf_thres, iou_thres, labels=lb, multi_label=True, agnostic=single_cls)
                t1 += time_synchronized() - t
                if opt.pack ==True:
                    if len(out[0] != 0):
                        key_frame +=1
                        old_out = [[]]
                        old_out[0] = out[0].clone().type(torch.cuda.LongTensor)
                        prev_out = out[0][:,:4]
                    else:
                        key_frame = 0  
                                 
            else:
                pack = True
                t = time_synchronized()
                # Since the inference time of the key frame is the longest, the canvas is formed here. However, conceptually, a canvas is formed after key frame inference.
                packed_img, where_old_image, where_new_image, groups_groups = canvas(img, out[0][:,:4], where_old_image, where_new_image, prev_out, opt.background)
                t_pack += time_synchronized() - t
                t = time_synchronized()
                if key_frame ==1:
                    box_minmax = [old_out[0][groups,:4] for groups in groups_groups]
                    old_box_flow_remember = []
                    for boxs in box_minmax:
                        if len(boxs) !=0:
                            old_box_flow_remember.append(torch.cat([torch.min(boxs[:,:2], dim=0, keepdim=True)[0],torch.max(boxs[:,2:], dim=0, keepdim=True)[0]], dim=1).type(torch.cuda.LongTensor))
                        else:
                            old_box_flow_remember.append(None)
                t3 += time_synchronized() - t
                # Run model
                # t_pack = time_synchronized()
                packed_img = F.pad(packed_img, (0,max(0,64-packed_img.size(3)),0,max(0,64- packed_img.size(2))), "constant", 0)
                
                all_packed_img.append(packed_img.size(2)*packed_img.size(3))
                t = time_synchronized()
                out, train_out = model(packed_img, augment=augment)  # inference and training outputs
                t0 += time_synchronized() - t

                # Compute loss
                if compute_loss:
                    loss += compute_loss([x.float() for x in train_out], targets)[1][:3]  # box, obj, cls

                # Run NMS
                targets[:, 2:] *= torch.Tensor([width, height, width, height]).to(device)  # to pixels
                lb = [targets[targets[:, 0] == i, 1:] for i in range(nb)] if save_hybrid else []  # for autolabelling
                
                t = time_synchronized()
                out = non_max_suppression(out, conf_thres, iou_thres, labels=lb, multi_label=True, agnostic=single_cls)
                t1 += time_synchronized() - t
                
                if len(out[0]) != 0:
                    if where_new_image == None:
                        t = time_synchronized()
                        out[0][:,:2] += where_old_image[:,:2]
                        out[0][:,2:4] += where_old_image[:,:2]
                        argmax = torch.zeros(len(out[0])).type(torch.cuda.LongTensor)
                        t2 += time_synchronized() - t 
                    else:
                        t = time_synchronized()
                        pack_box_iou = box_center(out[0][:,:4], where_new_image)
                        argmax = torch.argmax(pack_box_iou, dim=1)

                        check_for_edge = out[0][:,:4] - where_new_image[argmax,:4]
                        edge_thres = max(img.size(3), img.size(2))//100
                        index_for_chips_edge = (check_for_edge[:,0:1]<edge_thres)*1+(check_for_edge[:,1:2]<edge_thres)*1+(check_for_edge[:,2:3]>-edge_thres)*1+(check_for_edge[:,3:4]>-edge_thres)*1<=0
                        index_for_chips_edge = index_for_chips_edge

                        out[0][:,2:4] -= where_new_image[argmax,:2]
                        out[0][:,:2] -= where_new_image[argmax,:2]
                        out[0][:,:2] += where_old_image[argmax,:2]
                        out[0][:,2:4] += where_old_image[argmax,:2]

                        index_for_img_edge = 1*(out[0][:,0:1]<edge_thres)+1*(out[0][:,1:2]<edge_thres)+1*(out[0][:,2:3]>img.size(3)-edge_thres)+1*(out[0][:,3:4]>img.size(2)-edge_thres)>0
                        index_for_img_edge = index_for_img_edge

                        index_for_edge = index_for_chips_edge+index_for_img_edge>0

                        index_for_edge = index_for_edge.nonzero(as_tuple=True)[0]
                        out = torch.index_select(out[0], 0, index_for_edge).unsqueeze(0)


                        out = [out[0][torchvision.ops.nms(out[0][:,:4], out[0][:,4], 0.65)]]

                        if len(out[0]) != 0:
                            pack_box_iou = box_center(out[0][:,:4], where_old_image)
                            argmax = torch.argmax(pack_box_iou, dim=1)

                            t2 += time_synchronized() - t    
                            t = time_synchronized()

                            box_minmax = [out[0][argmax==k,:4].clone().type(torch.cuda.LongTensor) for k in range(len(where_old_image))]
                            box_flow_remember = []
                            for boxs in box_minmax:
                                if len(boxs) !=0:
                                    box_flow_remember.append(torch.cat([torch.min(boxs[:,:2], dim=0, keepdim=True)[0],torch.max(boxs[:,2:], dim=0, keepdim=True)[0]], dim=1))
                                else:
                                    box_flow_remember.append(None)
                            limit = max(img.size(3), img.size(2))//20
                            for index, (box_flow, old_box_flow) in enumerate(zip(box_flow_remember, old_box_flow_remember)):
                                if box_flow is None and old_box_flow is not None: 
                                    box_flow_remember[index] = old_box_flow
                                if box_flow is None or old_box_flow is None:
                                    continue
                                x_min_flow, x_max_flow, y_min_flow, y_max_flow = box_flow[0,0]-old_box_flow[0,0], box_flow[0,2]-old_box_flow[0,2], box_flow[0,1]-old_box_flow[0,1], box_flow[0,3]-old_box_flow[0,3]
                                x_min_flow, x_max_flow, y_min_flow, y_max_flow = x_min_flow.item(), x_max_flow.item(), y_min_flow.item(), y_max_flow.item()
                                if x_min_flow*x_max_flow>0 and abs(x_min_flow)<limit and abs(x_max_flow)<limit:
                                    some_to_add = [x_min_flow,x_max_flow][np.argmin([abs(x_min_flow),abs(x_max_flow)])]
                                    if some_to_add<0:
                                        if where_old_image[index,0]+some_to_add < 0:
                                            some_to_add = -where_old_image[index,0]
                                    else:
                                        if where_old_image[index,2]+some_to_add > img.size(3):
                                            some_to_add = img.size(3)-where_old_image[index,2]
                                    where_old_image[index,[0,2]] += some_to_add
                                if y_min_flow*y_max_flow>0 and abs(y_min_flow)<limit and abs(y_max_flow)<limit:
                                    some_to_add = [y_min_flow,y_max_flow][np.argmin([abs(y_min_flow),abs(y_max_flow)])]
                                    if some_to_add<0:
                                        if where_old_image[index,1]+some_to_add < 0:
                                            some_to_add = -where_old_image[index,1]
                                    else:
                                        if where_old_image[index,3]+some_to_add > img.size(2):
                                            some_to_add = img.size(2)-where_old_image[index,3]
                                    where_old_image[index,[1,3]] += some_to_add

                            old_box_flow_remember = box_flow_remember
                            t3 += time_synchronized() - t

                        else:
                            t2 += time_synchronized() - t

                    prev_out = out[0][:,:4]
                if len(out[0]) != 0:
                    key_frame +=1             
                else:
                    key_frame = 0 
                    
                if key_frame ==duration:
                    key_frame=0 


        # Statistics per image
        for si, pred in enumerate(out):
            labels = targets[targets[:, 0] == si, 1:]
            nl = len(labels)
            tcls = labels[:, 0].tolist() if nl else []  # target class
            path = Path(paths[si])
            seen += 1

            if len(pred) == 0:
                if nl:
                    stats.append((torch.zeros(0, niou, dtype=torch.bool), torch.Tensor(), torch.Tensor(), tcls))
                continue

            # Predictions
            if single_cls:
                pred[:, 5] = 0
            if opt.dataset_name =='AUAIR':
                pred[:, 5] = AUAIR_change_cls(pred[:, 5])
            if opt.dataset_name =='UAVDT':
                pred[:, 5] = UAVDT_change_cls(pred[:, 5])
    
            pred, out[si] = pred[pred[:, 5]!=-1], out[si][pred[:, 5]!=-1]
        
            if len(pred) == 0:
                if nl:
                    stats.append((torch.zeros(0, niou, dtype=torch.bool), torch.Tensor(), torch.Tensor(), tcls))
                continue    
    
            predn = pred.clone()
            scale_coords(img[si].shape[1:], predn[:, :4], shapes[si][0], shapes[si][1])  # native-space pred

            # Append to text file
            if save_txt:
                gn = torch.tensor(shapes[si][0])[[1, 0, 1, 0]]  # normalization gain whwh
                for *xyxy, conf, cls in predn.tolist():
                    xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
                    line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format
                    with open(save_dir / 'labels' / (path.stem + '.txt'), 'a') as f:
                        f.write(('%g ' * len(line)).rstrip() % line + '\n')

            # W&B logging - Media Panel Plots
            if len(wandb_images) < log_imgs and wandb_logger.current_epoch > 0:  # Check for test operation
                if wandb_logger.current_epoch % wandb_logger.bbox_interval == 0:
                    box_data = [{"position": {"minX": xyxy[0], "minY": xyxy[1], "maxX": xyxy[2], "maxY": xyxy[3]},
                                 "class_id": int(cls),
                                 "box_caption": "%s %.3f" % (names[cls], conf),
                                 "scores": {"class_score": conf},
                                 "domain": "pixel"} for *xyxy, conf, cls in pred.tolist()]
                    boxes = {"predictions": {"box_data": box_data, "class_labels": names}}  # inference-space
                    wandb_images.append(wandb_logger.wandb.Image(img[si], boxes=boxes, caption=path.name))
            wandb_logger.log_training_progress(predn, path, names) if wandb_logger and wandb_logger.wandb_run else None

            # Append to pycocotools JSON dictionary
            if save_json:
                # [{"image_id": 42, "category_id": 18, "bbox": [258.15, 41.29, 348.26, 243.78], "score": 0.236}, ...
                image_id = int(path.stem) if path.stem.isnumeric() else path.stem
                box = xyxy2xywh(predn[:, :4])  # xywh
                box[:, :2] -= box[:, 2:] / 2  # xy center to top-left corner
                for p, b in zip(pred.tolist(), box.tolist()):
                    jdict.append({'image_id': image_id,
                                  'category_id': coco91class[int(p[5])] if is_coco else int(p[5]),
                                  'bbox': [round(x, 3) for x in b],
                                  'score': round(p[4], 5)})

            # Assign all predictions as incorrect
            correct = torch.zeros(pred.shape[0], niou, dtype=torch.bool, device=device)
            if nl:
                detected = []  # target indices
                tcls_tensor = labels[:, 0]

                # target boxes
                tbox = xywh2xyxy(labels[:, 1:5])
                scale_coords(img[si].shape[1:], tbox, shapes[si][0], shapes[si][1])  # native-space labels
                if plots:
                    confusion_matrix.process_batch(predn, torch.cat((labels[:, 0:1], tbox), 1))

                # Per target class
                for cls in torch.unique(tcls_tensor):
                    ti = (cls == tcls_tensor).nonzero(as_tuple=False).view(-1)  # target indices
                    pi = (cls == pred[:, 5]).nonzero(as_tuple=False).view(-1)  # prediction indices

                    # Search for detections
                    if pi.shape[0]:
                        # Prediction to target ious
                        ious, i = box_iou(predn[pi, :4], tbox[ti]).max(1)  # best ious, indices

                        # Append detections
                        detected_set = set()
                        for j in (ious > iouv[0]).nonzero(as_tuple=False):
                            d = ti[i[j]]  # detected target
                            if d.item() not in detected_set:
                                detected_set.add(d.item())
                                detected.append(d)
                                correct[pi[j]] = ious[j] > iouv  # iou_thres is 1xn
                                if len(detected) == nl:  # all targets already located in image
                                    break

            # Append statistics (correct, conf, pcls, tcls)
            stats.append((correct.cpu(), pred[:, 4].cpu(), pred[:, 5].cpu(), tcls))

        # Plot images
        if plots and batch_i < opt.saved_frames:
            n = '0'*(4-len(str(batch_i)))+str(batch_i)
            f = save_dir / f'test_batch{n}_labels.jpg'  # labels
            Thread(target=plot_images, args=(img, targets, paths, f, names), daemon=True).start()
            f = save_dir / f'test_batch{n}_pred.jpg'  # predictions
            Thread(target=plot_images, args=(img, output_to_target(out), paths, f, names), daemon=True).start()
            if pack == True:
                f = save_dir / f'packed_frames/pack_image{n}.jpg'  # predictions
                Thread(target=plot_images, args=(packed_img, output_to_target([out[0]*0]), paths, f, names), daemon=True).start()               
    # Compute statistics
    stats = [np.concatenate(x, 0) for x in zip(*stats)]  # to numpy
    if len(stats) and stats[0].any():
        p, r, ap, f1, ap_class = ap_per_class(*stats, plot=plots, save_dir=save_dir, names=names)
        ap50, ap = ap[:, 0], ap.mean(1)  # AP@0.5, AP@0.5:0.95
        mp, mr, map50, map = p.mean(), r.mean(), ap50.mean(), ap.mean()
        nt = np.bincount(stats[3].astype(np.int64), minlength=nc)  # number of targets per class
    else:
        nt = torch.zeros(1)

    # Print results
    pf = '%20s' + '%12i' * 2 + '%12.3g' * 4  # print format
    print(pf % ('all', seen, nt.sum(), mp, mr, map50, map))

    # Print results per class
    if (verbose or (nc < 50 and not training)) and nc > 1 and len(stats):
        for i, c in enumerate(ap_class):
            print(pf % (names[c], seen, nt[c], p[i], r[i], ap50[i], ap[i]))

    # Print speeds
    t = tuple(x / seen * 1E3 for x in (t_pack, t0, t1,t2,t3, t_pack+t0 + t1+t2+t3)) + (imgsz, imgsz, batch_size)  # tuple

    
    if not training:
        print('Speed: %.1f/%.1f/%.1f/%.1f/%.1f/%.1f ms t_pack/inference/NMS/Back_to_origianl/PAd_Track/total per %gx%g image at batch-size %g' % t)

    # Saved in excel
    if opt.pack == False:
        load_wb = load_workbook(f'excel_result/{opt.dataset_name}_results_baseline.xlsx', data_only = True)
        all_size_img= np.mean(all_anchor_img)
    else:
        load_wb = load_workbook(f'excel_result/{opt.dataset_name}_results_DynaPP.xlsx', data_only = True)
        all_size_img = np.mean(all_packed_img+all_anchor_img)
    all_size_img = np.sqrt(all_size_img)
    load_ws = load_wb['Sheet']
    all_values = []
    
    for row in load_ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)
    wr = [data['val'].split('/')[-2],t[0],t[1],t[2],t[3],t[4],t[5],map50, map, all_size_img]
    all_values.append(wr)
    
    write_wb =Workbook()
    write_ws = write_wb.active
    for i in range(len(all_values)):
        write_ws.append(list(all_values[i]))
    if opt.pack == False:
        write_wb.save(f'excel_result/{opt.dataset_name}_results_baseline.xlsx')
    else:
        write_wb.save(f'excel_result/{opt.dataset_name}_results_DynaPP.xlsx')
    
    # Plots
    if plots:
        confusion_matrix.plot(save_dir=save_dir, names=list(names.values()))
        if wandb_logger and wandb_logger.wandb:
            val_batches = [wandb_logger.wandb.Image(str(f), caption=f.name) for f in sorted(save_dir.glob('test*.jpg'))]
            wandb_logger.log({"Validation": val_batches})
    if wandb_images:
        wandb_logger.log({"Bounding Box Debugger/Images": wandb_images})

    # Save JSON
    if save_json and len(jdict):
        w = Path(weights[0] if isinstance(weights, list) else weights).stem if weights is not None else ''  # weights
        anno_json = '../coco/annotations/instances_val2017.json'  # annotations json
        pred_json = str(save_dir / f"{w}_predictions.json")  # predictions json
        print('\nEvaluating pycocotools mAP... saving %s...' % pred_json)
        with open(pred_json, 'w') as f:
            json.dump(jdict, f)

        try:  # https://github.com/cocodataset/cocoapi/blob/master/PythonAPI/pycocoEvalDemo.ipynb
            from pycocotools.coco import COCO
            from pycocotools.cocoeval import COCOeval

            anno = COCO(anno_json)  # init annotations api
            pred = anno.loadRes(pred_json)  # init predictions api
            eval = COCOeval(anno, pred, 'bbox')
            if is_coco:
                eval.params.imgIds = [int(Path(x).stem) for x in dataloader.dataset.img_files]  # image IDs to evaluate
            eval.evaluate()
            eval.accumulate()
            eval.summarize()
            map, map50 = eval.stats[:2]  # update results (mAP@0.5:0.95, mAP@0.5)
        except Exception as e:
            print(f'pycocotools unable to run: {e}')

    # Return results
    model.float()  # for training
    if not training:
        s = f"\n{len(list(save_dir.glob('labels/*.txt')))} labels saved to {save_dir / 'labels'}" if save_txt else ''
        print(f"Results saved to {save_dir}{s}")
    maps = np.zeros(nc) + map
    for i, c in enumerate(ap_class):
        maps[c] = ap[i]
    return (mp, mr, map50, map, *(loss.cpu() / len(dataloader)).tolist()), maps, t


if __name__ == '__main__':
    parser = argparse.ArgumentParser(prog='test.py')
    parser.add_argument('--weights', nargs='+', type=str, default='yolov5s.pt', help='model.pt path(s)')
    parser.add_argument('--data', type=str, default='data/coco128.yaml', help='*.data path')
    parser.add_argument('--batch-size', type=int, default=32, help='size of each image batch')
    parser.add_argument('--img_size', type=int, default=640, help='inference size (pixels)')
    parser.add_argument('--conf-thres', type=float, default=0.001, help='object confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.6, help='IOU threshold for NMS')
    parser.add_argument('--task', default='val', help='train, val, test, speed or study')
    parser.add_argument('--device', default='', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
    parser.add_argument('--single-cls', action='store_true', help='treat as single-class dataset')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    parser.add_argument('--verbose', action='store_true', help='report mAP by class')
    parser.add_argument('--save-txt', action='store_true', help='save results to *.txt')
    parser.add_argument('--save-hybrid', action='store_true', help='save label+prediction hybrid results to *.txt')
    parser.add_argument('--save-conf', action='store_true', help='save confidences in --save-txt labels')
    parser.add_argument('--save-json', action='store_true', help='save a cocoapi-compatible JSON results file')
    parser.add_argument('--project', default='runs/test', help='save to project/name')
    parser.add_argument('--name', default='exp', help='save to project/name')
    parser.add_argument('--exist-ok', action='store_true', help='existing project/name ok, do not increment')
    parser.add_argument('--pack', action='store_true', help='Use packing')
    parser.add_argument('--dataset_name', type=str, default='None', help='dataset name')
    parser.add_argument('--duration', type=int, default=5, help='key frame duration length')
    parser.add_argument('--background', type=float, default=0.3, help='background d%')
    parser.add_argument('--saved_frames', type=int, default=100, help='number of the saved frames')


    opt = parser.parse_args()
    opt.save_json |= opt.data.endswith('coco.yaml')
    opt.data = check_file(opt.data)  # check file
    print(opt)
    check_requirements(exclude=('tensorboard', 'pycocotools', 'thop'))
    if opt.task in ('train', 'val', 'test'):  # run normally
        test(opt.data,
             opt.weights,
             opt.batch_size,
             opt.img_size,
             opt.conf_thres,
             opt.iou_thres,
             opt.save_json,
             opt.single_cls,
             opt.augment,
             opt.verbose,
             save_txt=opt.save_txt | opt.save_hybrid,
             save_hybrid=opt.save_hybrid,
             save_conf=opt.save_conf,
             opt=opt
             )

 
    elif opt.task == 'speed':  # speed benchmarks
        for w in opt.weights:
            test(opt.data, w, opt.batch_size, opt.img_size, 0.25, 0.45, save_json=False, plots=False, opt=opt)

    elif opt.task == 'study':  # run over a range of settings and save/plot
        # python test.py --task study --data coco.yaml --iou 0.7 --weights yolov5s.pt yolov5m.pt yolov5l.pt yolov5x.pt
        x = list(range(256, 1536 + 128, 128))  # x axis (image sizes)
        for w in opt.weights:
            f = f'study_{Path(opt.data).stem}_{Path(w).stem}.txt'  # filename to save to
            y = []  # y axis
            for i in x:  # img-size
                print(f'\nRunning {f} point {i}...')
                r, _, t = test(opt.data, w, opt.batch_size, i, opt.conf_thres, opt.iou_thres, opt.save_json,
                               plots=False, opt=opt)
                y.append(r + t)  # results and times
            np.savetxt(f, y, fmt='%10.4g')  # save
        os.system('zip -r study.zip study_*.txt')
        plot_study_txt(x=x)  # plot
